import json
from pathlib import Path

from openpyxl import load_workbook


# ==========================================================
# DATASET ROOT
# ==========================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

DATASET_ROOT = PROJECT_ROOT / "dataset"

REFERENCES_DIR = DATASET_ROOT / "references"

REFERENCE_DATASET = (
    REFERENCES_DIR
    / "leetcode_1_to_10_reference_dataset_minimal.xlsx"
)

RUBRICS_DIR = DATASET_ROOT / "rubrics"


# ==========================================================
# HELPERS
# ==========================================================

def _load_json(path: Path) -> dict:
    """
    Load a JSON file and return its contents.
    """

    if not path.exists():
        raise FileNotFoundError(
            f"Dataset file not found: {path}"
        )

    try:
        with path.open(
            "r",
            encoding="utf-8"
        ) as file:
            data = json.load(file)

    except json.JSONDecodeError as error:
        raise ValueError(
            f"Invalid JSON in dataset file: {path}"
        ) from error

    if not isinstance(data, dict):
        raise TypeError(
            f"Dataset file must contain a JSON object: {path}"
        )

    return data


def _get_problem_id(problem: dict) -> str:
    """
    Extract the dataset problem ID from the problem object.
    """

    if not isinstance(problem, dict):
        raise TypeError(
            "problem must be a dictionary."
        )

    problem_id = (
        problem.get("problem_id")
        or problem.get("id")
    )

    if problem_id:
        return str(
            problem_id
        ).strip()

    raise ValueError(
        "Problem does not contain 'id' or 'problem_id'."
    )


# ==========================================================
# REFERENCE DATASET
# ==========================================================

def load_reference_dataset() -> dict:
    """
    Load all reference solutions from the Excel dataset.

    Returns:

        {
            "P001": {
                "problem_id": "P001",
                "problem_title": "...",
                "reference_solutions": [...]
            },
            ...
        }
    """

    if not REFERENCE_DATASET.exists():
        raise FileNotFoundError(
            f"Reference dataset not found: "
            f"{REFERENCE_DATASET}"
        )

    workbook = load_workbook(
        REFERENCE_DATASET,
        read_only=True,
        data_only=True
    )

    try:

        if "Reference Solutions" not in workbook.sheetnames:
            raise ValueError(
                "Excel dataset does not contain the "
                "'Reference Solutions' sheet."
            )

        sheet = workbook[
            "Reference Solutions"
        ]

        rows = sheet.iter_rows(
            values_only=True
        )

        headers = next(
            rows,
            None
        )

        if not headers:
            raise ValueError(
                "Reference dataset is empty."
            )

        headers = [
            str(header).strip()
            if header is not None
            else ""
            for header in headers
        ]

        dataset = {}

        for row in rows:

            if not any(
                value is not None
                for value in row
            ):
                continue

            reference = {}

            for index, header in enumerate(headers):

                if not header:
                    continue

                value = (
                    row[index]
                    if index < len(row)
                    else None
                )

                reference[
                    header
                ] = value

            problem_id = reference.get(
                "Problem ID"
            )

            if problem_id is None:
                continue

            problem_id = str(
                problem_id
            ).strip()

            reference_id = reference.get(
                "Reference ID"
            )

            if reference_id is not None:
                reference[
                    "Reference ID"
                ] = str(
                    reference_id
                ).strip()

            if problem_id not in dataset:

                dataset[
                    problem_id
                ] = {
                    "problem_id": problem_id,
                    "problem_title": reference.get(
                        "Problem Title"
                    ),
                    "reference_solutions": []
                }

            dataset[
                problem_id
            ][
                "reference_solutions"
            ].append(
                reference
            )

        return dataset

    finally:

        workbook.close()


# ==========================================================
# REFERENCE SOLUTIONS
# ==========================================================

def load_reference_solution(
    problem: dict
) -> list[dict]:
    """
    Load all reference solutions for a problem.

    A problem may have multiple reference solutions.

    Returns a list containing all references for that problem.
    """

    problem_id = _get_problem_id(
        problem
    )

    dataset = load_reference_dataset()

    problem_data = dataset.get(
        problem_id
    )

    if problem_data is None:
        raise FileNotFoundError(
            f"No reference solutions found for "
            f"problem: {problem_id}"
        )

    references = problem_data.get(
        "reference_solutions",
        []
    )

    if not references:
        raise ValueError(
            f"Problem {problem_id} has no reference solutions."
        )

    return references


# ==========================================================
# RUBRIC
# ==========================================================

def load_rubric(
    problem: dict
) -> dict:
    """
    Load the evaluation rubric for a problem.
    """

    problem_id = _get_problem_id(
        problem
    )

    path = (
        RUBRICS_DIR
        / f"{problem_id}_rubric.json"
    )

    return _load_json(
        path
    )


# ==========================================================
# EVALUATION CONTEXT
# ==========================================================

def load_evaluation_context(
    problem: dict
) -> tuple[list[dict], dict]:
    """
    Load all reference solutions and the rubric
    for the given problem.

    Returns:

        (
            reference_solutions,
            rubric
        )
    """

    reference_solutions = (
        load_reference_solution(
            problem
        )
    )

    rubric = load_rubric(
        problem
    )

    return (
        reference_solutions,
        rubric
    )